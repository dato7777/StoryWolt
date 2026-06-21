"""
Core commission calculation engine for Wolt net income.

This module loads commission rates from offers_commission.xlsx and applies them
to rows parsed from a Wolt itemsSold CSV export uploaded by the merchant.

Business rule (per Story Phone / Wolt Israel):
    commission_before_vat = selling_price_incl_vat * (commission_percent / 100)
    total_wolt_commission = commission_before_vat * VAT_MULTIPLIER
    net_income = selling_price_incl_vat - total_wolt_commission - (product_self_cost * quantity)

The commission percent itself does not include VAT; VAT is added on top of
Wolt's fee invoice (18% in Israel).
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import asdict, dataclass, replace
from datetime import date
from pathlib import Path
from typing import Any

# Israeli VAT rate applied to Wolt's commission fee (not to the product price).
VAT_MULTIPLIER = 1.18

# Expected columns in the uploaded Wolt itemsSold export.
REQUIRED_CSV_COLUMNS = {"Item name", "Total", "Quantity"}
OPTIONAL_CSV_COLUMNS = {"Merchant SKU", "GTIN", "POS ID"}

# Expected columns in the uploaded Wolt orderNumbers export.
REQUIRED_ORDER_COLUMNS = {"Order number", "Delivery status", "Items", "Price"}

# Column name in offers_commission.xlsx for merchant purchase cost per unit (incl. VAT).
SELF_COST_COLUMN = "Product Self cost (including VAT)"

# Parses "3x Product name 119 ILS" fragments inside orderNumbers Items column.
ITEM_PATTERN = re.compile(r"(\d+)x\s+(.+?)\s+([\d.]+)\s+ILS")

AD_CAMPAIGN_BETWEEN_PATTERN = re.compile(
    r"between (\d{4}-\d{2}-\d{2}) - (\d{4}-\d{2}-\d{2})"
)
AD_CAMPAIGN_ON_PATTERN = re.compile(r"on (\d{4}-\d{2}-\d{2})\b")


def normalize_product_name(name: str) -> str:
    """
    Normalize product names so catalog and sales exports can be matched reliably.

    - Lowercase
    - Collapse whitespace
    - Treat pipe separators as spaces (common in Wolt Hebrew titles)
    """
    return re.sub(r"\s+", " ", name.lower().replace("|", " ").strip())


@dataclass
class CommissionOffer:
    """Single product row from offers_commission.xlsx."""

    merchant_sku: str
    name: str
    price: float | None
    commission_percent: float | None
    self_cost: float = 0.0


@dataclass
class SalesRow:
    """Single aggregated sales line from itemsSold.csv."""

    merchant_sku: str
    item_name: str
    total_gross: float
    quantity: int


@dataclass
class CalculatedRow:
    """Output row shown in the dashboard after commission is applied."""

    item_name: str
    merchant_sku: str
    quantity: int
    list_price: float | None
    list_total: float | None
    sold_total: float
    gross_total: float
    commission_percent: float | None
    commission_before_vat: float
    commission_with_vat: float
    commission_with_vat_per_item: float
    product_self_cost: float
    net_income: float
    net_income_per_item: float
    status: str
    match_method: str
    allocated_ad_cost: float = 0.0
    net_income_after_ad_cost: float = 0.0
    net_income_after_ad_cost_per_item: float = 0.0


@dataclass
class CalculationSummary:
    """Period-level totals returned to the frontend summary cards."""

    row_count: int
    matched_count: int
    unmatched_count: int
    delivered_order_count: int
    rejected_order_count: int
    rejected_order_total: float
    total_gross: float
    total_list_value: float
    total_sold_value: float
    total_commission_before_vat: float
    total_commission_with_vat: float
    total_net_income: float
    total_product_self_cost: float
    # Populated when standardSummary / payment_details CSV is uploaded.
    wolt_summary_gross_goods: float | None = None
    wolt_summary_expenses_net: float | None = None
    wolt_summary_expenses_incl_vat: float | None = None
    wolt_summary_distribution_incl_vat: float | None = None
    wolt_summary_remunerations: float | None = None
    wolt_summary_self_billing_deductions_incl_vat: float | None = None
    wolt_summary_self_billing_negative_incl_vat: float | None = None
    wolt_summary_payout: float | None = None
    wolt_summary_net_income: float | None = None
    wolt_summary_ad_campaigns_incl_vat: float | None = None
    wolt_summary_ad_campaigns_allocated_incl_vat: float | None = None
    wolt_summary_other_fees_incl_vat: float | None = None
    wolt_summary_distribution_gap_incl_vat: float | None = None
    per_item_expenses_excluded_incl_vat: float | None = None
    per_item_expenses_excluded_after_ads_incl_vat: float | None = None


@dataclass
class AdCampaignCharge:
    """One ad campaign line from WOLT INVOICE with a purchasable date window."""

    label: str
    amount_incl_vat: float
    start_date: date
    end_date: date


@dataclass
class OrderLineItem:
    """Single product line inside one delivered order."""

    item_name: str
    merchant_sku: str
    quantity: int
    line_gross: float
    list_price: float | None
    commission_percent: float | None
    commission_before_vat: float
    commission_with_vat: float
    commission_with_vat_per_item: float
    product_self_cost: float
    net_income: float
    net_income_per_item: float
    status: str
    match_method: str
    allocated_ad_cost: float = 0.0
    net_income_after_ad_cost: float = 0.0
    net_income_after_ad_cost_per_item: float = 0.0


@dataclass
class CalculatedOrder:
    """Delivered order with per-item breakdown and order-level totals."""

    order_number: str
    order_placed: str
    delivery_time: str
    delivery_status: str
    order_gross: float
    commission_before_vat: float
    commission_with_vat: float
    net_income: float
    items: list[OrderLineItem]
    allocated_ad_cost: float = 0.0
    net_income_after_ad_cost: float = 0.0


@dataclass
class InvoiceStep:
    """One row in the invoice reconciliation waterfall (display only)."""

    id: str
    label: str
    label_he: str
    amount: float
    running_total: float
    step_type: str
    note: str
    phase: str


@dataclass
class InvoicePhase:
    """Grouped waterfall section shown as expandable accordion in the UI."""

    id: str
    title: str
    subtitle: str
    steps: list[InvoiceStep]


@dataclass
class InvoiceReconciliation:
    """
    Display-only bridge from Wolt self-billing invoice to order totals.

    Does not change per-item commission or net income calculations.
    """

    source: str
    gross_goods_sold: float | None
    merchant_discounts: float | None
    net_sold_from_invoice: float | None
    net_sold_from_orders: float
    orders_match_invoice: bool | None
    remunerations: float | None
    net_after_remunerations: float | None
    wolt_distribution_fees: float
    net_income_after_wolt: float
    payout_amount: float | None
    total_wolt_invoice: float | None
    payout_gap_from_app_net: float | None
    steps: list[InvoiceStep]
    phases: list[InvoicePhase]


def load_offers_from_xlsx(path: Path) -> dict[str, CommissionOffer]:
    """
    Load commission lookup table from offers_commission.xlsx.

    Returns a dict keyed by normalized product name. SKU index is built separately
    in build_offer_indexes().
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read offers_commission.xlsx") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    workbook.close()

    if not rows:
        return {}

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    name_idx = headers.index("name")
    sku_idx = headers.index("merchant_sku") if "merchant_sku" in headers else None
    price_idx = headers.index("price") if "price" in headers else None
    commission_idx = headers.index("commission_home_delivery")

    offers: dict[str, CommissionOffer] = {}
    for row in rows[1:]:
        if not row or len(row) <= name_idx or not row[name_idx]:
            continue

        name = str(row[name_idx]).strip()
        sku = str(row[sku_idx]).strip() if sku_idx is not None and row[sku_idx] else ""
        price_raw = row[price_idx] if price_idx is not None else None
        commission_raw = row[commission_idx] if len(row) > commission_idx else None

        price = float(price_raw) if price_raw not in (None, "") else None
        commission = (
            float(commission_raw) if commission_raw not in (None, "") else None
        )
        self_cost_idx = headers.index(SELF_COST_COLUMN) if SELF_COST_COLUMN in headers else None
        self_cost_raw = (
            row[self_cost_idx]
            if self_cost_idx is not None and len(row) > self_cost_idx
            else None
        )
        self_cost = float(self_cost_raw) if self_cost_raw not in (None, "") else 0.0

        offers[normalize_product_name(name)] = CommissionOffer(
            merchant_sku=sku,
            name=name,
            price=price,
            commission_percent=commission,
            self_cost=round(self_cost, 2),
        )

    return offers


def build_offer_indexes(
    offers_by_name: dict[str, CommissionOffer],
) -> tuple[dict[str, CommissionOffer], dict[str, CommissionOffer]]:
    """
    Build two lookup indexes:
    - by normalized name (exact after normalization)
    - by merchant SKU string
    """
    by_sku: dict[str, CommissionOffer] = {}
    for offer in offers_by_name.values():
        if offer.merchant_sku:
            by_sku[offer.merchant_sku.strip()] = offer
    return offers_by_name, by_sku


def parse_items_sold_csv(csv_text: str) -> list[SalesRow]:
    """
    Parse the uploaded Wolt itemsSold CSV export.

    Recommended upload format for merchants:
        sales 0106-1506 itemsSold.csv  (Wolt export: aggregated product sales)

    Required headers: Item name, Total, Quantity
    Optional headers: Merchant SKU, GTIN, POS ID
    """
    reader = csv.DictReader(io.StringIO(csv_text))
    if not reader.fieldnames:
        raise ValueError("CSV file is empty or missing a header row.")

    fieldnames = {name.strip() for name in reader.fieldnames}
    missing = REQUIRED_CSV_COLUMNS - fieldnames
    if missing:
        raise ValueError(
            "Invalid file type. Upload Wolt itemsSold CSV with columns: "
            + ", ".join(sorted(REQUIRED_CSV_COLUMNS))
            + f". Missing: {', '.join(sorted(missing))}"
        )

    rows: list[SalesRow] = []
    for line_number, record in enumerate(reader, start=2):
        item_name = (record.get("Item name") or "").strip()
        if not item_name:
            continue

        try:
            total_gross = float(str(record.get("Total", "0")).replace(",", "").strip())
            quantity = int(float(str(record.get("Quantity", "0")).replace(",", "").strip()))
        except ValueError as exc:
            raise ValueError(f"Invalid number on CSV line {line_number}.") from exc

        rows.append(
            SalesRow(
                merchant_sku=(record.get("Merchant SKU") or "").strip(),
                item_name=item_name,
                total_gross=total_gross,
                quantity=quantity,
            )
        )

    if not rows:
        raise ValueError("No product rows found in uploaded CSV.")

    return rows


def lookup_offer_by_name(
    item_name: str,
    merchant_sku: str,
    offers_by_name: dict[str, CommissionOffer],
    offers_by_sku: dict[str, CommissionOffer],
) -> tuple[CommissionOffer | None, str]:
    """Resolve catalog offer for an item name and optional merchant SKU."""
    if merchant_sku and merchant_sku in offers_by_sku:
        return offers_by_sku[merchant_sku], "merchant_sku"

    normalized = normalize_product_name(item_name)
    if normalized in offers_by_name:
        return offers_by_name[normalized], "exact_name"

    for catalog_name, offer in offers_by_name.items():
        if normalized == catalog_name or normalized in catalog_name or catalog_name in normalized:
            return offer, "fuzzy_name"

    return None, "not_found"


def lookup_offer(
    sales_row: SalesRow,
    offers_by_name: dict[str, CommissionOffer],
    offers_by_sku: dict[str, CommissionOffer],
) -> tuple[CommissionOffer | None, str]:
    """Resolve the catalog offer row for a sold item from itemsSold.csv."""
    return lookup_offer_by_name(
        sales_row.item_name, sales_row.merchant_sku, offers_by_name, offers_by_sku
    )


def lookup_commission_percent(
    sales_row: SalesRow,
    offers_by_name: dict[str, CommissionOffer],
    offers_by_sku: dict[str, CommissionOffer],
) -> tuple[float | None, str]:
    """Return commission percent and match method for a sold item."""
    offer, match_method = lookup_offer(sales_row, offers_by_name, offers_by_sku)
    if offer is None:
        return None, match_method
    return offer.commission_percent, match_method


def commission_base_amount(sales_row: SalesRow) -> float:
    """
    Amount used for Wolt commission — matches Wolt invoices.

    Uses the actual sold total from itemsSold.csv (incl. VAT, after discounts),
    not the catalog list price from offers_commission.xlsx.
    """
    return round(sales_row.total_gross, 2)


def build_calculated_row(
    sales_row: SalesRow,
    offer: CommissionOffer | None,
    match_method: str,
) -> CalculatedRow:
    """Build one dashboard row with list price, sold total, commission, and net income."""
    sold_total = round(sales_row.total_gross, 2)
    list_price = round(offer.price, 2) if offer and offer.price is not None else None
    list_total = round(list_price * sales_row.quantity, 2) if list_price is not None else None
    commission_percent = offer.commission_percent if offer else None
    product_self_cost = unit_self_cost(offer)

    if commission_percent is None:
        net_income = calculate_net_income_totally(
            sold_total, 0.0, product_self_cost, sales_row.quantity
        )
        return CalculatedRow(
            item_name=sales_row.item_name,
            merchant_sku=sales_row.merchant_sku,
            quantity=sales_row.quantity,
            list_price=list_price,
            list_total=list_total,
            sold_total=sold_total,
            gross_total=sold_total,
            commission_percent=None,
            commission_before_vat=0.0,
            commission_with_vat=0.0,
            commission_with_vat_per_item=0.0,
            product_self_cost=product_self_cost,
            net_income=net_income,
            net_income_per_item=net_income_per_unit(net_income, sales_row.quantity),
            status="missing_commission",
            match_method=match_method,
        )

    commission_base = commission_base_amount(sales_row)
    before_vat, with_vat = calculate_commission_amounts(commission_base, commission_percent)
    net_income = calculate_net_income_totally(
        sold_total, with_vat, product_self_cost, sales_row.quantity
    )

    return CalculatedRow(
        item_name=sales_row.item_name,
        merchant_sku=sales_row.merchant_sku,
        quantity=sales_row.quantity,
        list_price=list_price,
        list_total=list_total,
        sold_total=sold_total,
        gross_total=sold_total,
        commission_percent=commission_percent,
        commission_before_vat=before_vat,
        commission_with_vat=with_vat,
        commission_with_vat_per_item=fee_per_item(with_vat, sales_row.quantity),
        product_self_cost=product_self_cost,
        net_income=net_income,
        net_income_per_item=net_income_per_unit(net_income, sales_row.quantity),
        status="ok",
        match_method=match_method,
    )


def calculate_commission_amounts(gross_total: float, commission_percent: float) -> tuple[float, float]:
    """
    Apply Story Phone Wolt commission formula for one product line.

    Returns:
        (commission_before_vat, commission_with_vat)
    """
    commission_before_vat = gross_total * (commission_percent / 100.0)
    commission_with_vat = commission_before_vat * VAT_MULTIPLIER
    return round(commission_before_vat, 2), round(commission_with_vat, 2)


def unit_self_cost(offer: CommissionOffer | None) -> float:
    """Per-unit merchant cost from offers sheet (incl. VAT); defaults to 0."""
    if offer is None:
        return 0.0
    return round(offer.self_cost, 2)


def net_income_per_unit(net_income: float, quantity: int) -> float:
    """Net income for a single unit after Wolt fees and product self cost."""
    if quantity <= 0:
        return 0.0
    return round(net_income / quantity, 2)


def fee_per_item(commission_with_vat: float, quantity: int) -> float:
    """Wolt fee (incl. VAT) for a single unit."""
    if quantity <= 0:
        return 0.0
    return round(commission_with_vat / quantity, 2)


def line_self_cost_total(product_self_cost: float, quantity: int) -> float:
    """Total merchant cost for a line (per-unit self cost × quantity)."""
    return round(product_self_cost * quantity, 2)


def calculate_net_income_totally(
    sold_total: float,
    commission_with_vat: float,
    product_self_cost: float,
    quantity: int,
) -> float:
    """Net income after Wolt fee and product self cost (incl. VAT)."""
    return round(
        sold_total - commission_with_vat - line_self_cost_total(product_self_cost, quantity),
        2,
    )


def calculate_net_income_report(
    sales_rows: list[SalesRow],
    offers_by_name: dict[str, CommissionOffer],
    offers_by_sku: dict[str, CommissionOffer],
) -> tuple[list[CalculatedRow], CalculationSummary]:
    """
    Main orchestration function used by the serverless API.

    For each sold item:
    - find commission %
    - compute Wolt fee (with VAT on fee)
    - compute net income after Wolt taxation
    """
    calculated_rows: list[CalculatedRow] = []
    matched_count = 0

    for sales_row in sales_rows:
        offer, match_method = lookup_offer(sales_row, offers_by_name, offers_by_sku)
        row = build_calculated_row(sales_row, offer, match_method)
        if row.status == "ok":
            matched_count += 1
        calculated_rows.append(row)

    calculated_rows.sort(key=lambda row: row.sold_total, reverse=True)

    total_sold = sum(row.sold_total for row in calculated_rows)
    total_list = sum(row.list_total if row.list_total is not None else row.sold_total for row in calculated_rows)
    total_before_vat = sum(row.commission_before_vat for row in calculated_rows)
    total_with_vat = sum(row.commission_with_vat for row in calculated_rows)
    total_net = sum(row.net_income for row in calculated_rows)
    total_self_cost = sum(row.product_self_cost * row.quantity for row in calculated_rows)

    summary = CalculationSummary(
        row_count=len(calculated_rows),
        matched_count=matched_count,
        unmatched_count=len(calculated_rows) - matched_count,
        delivered_order_count=0,
        rejected_order_count=0,
        rejected_order_total=0.0,
        total_gross=round(total_sold, 2),
        total_list_value=round(total_list, 2),
        total_sold_value=round(total_sold, 2),
        total_commission_before_vat=round(total_before_vat, 2),
        total_commission_with_vat=round(total_with_vat, 2),
        total_net_income=round(total_net, 2),
        total_product_self_cost=round(total_self_cost, 2),
    )

    return calculated_rows, summary


def parse_order_numbers_csv(csv_text: str) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    """
    Parse Wolt orderNumbers CSV and split delivered vs rejected orders.

    Rejected orders are excluded from commission calculations but counted in summary.
    """
    reader = csv.DictReader(io.StringIO(csv_text))
    if not reader.fieldnames:
        raise ValueError("Order numbers CSV is empty or missing a header row.")

    fieldnames = {name.strip() for name in reader.fieldnames}
    missing = REQUIRED_ORDER_COLUMNS - fieldnames
    if missing:
        raise ValueError(
            "Invalid orderNumbers file. Required columns: "
            + ", ".join(sorted(REQUIRED_ORDER_COLUMNS))
            + f". Missing: {', '.join(sorted(missing))}"
        )

    delivered: list[dict[str, str]] = []
    rejected: list[dict[str, str]] = []
    for record in reader:
        status = (record.get("Delivery status") or "").strip().lower()
        if status == "delivered":
            delivered.append(record)
        elif status == "rejected":
            rejected.append(record)

    if not delivered:
        raise ValueError("No delivered orders found in orderNumbers CSV.")

    return delivered, rejected


def build_sku_map_from_items_sold(items_sold_rows: list[SalesRow]) -> dict[str, str]:
    """Map normalized product name to merchant SKU using optional itemsSold upload."""
    sku_map: dict[str, str] = {}
    for row in items_sold_rows:
        if row.merchant_sku:
            sku_map[normalize_product_name(row.item_name)] = row.merchant_sku
    return sku_map


def build_order_line_item(
    item_name: str,
    quantity: int,
    unit_price: float,
    merchant_sku: str,
    offers_by_name: dict[str, CommissionOffer],
    offers_by_sku: dict[str, CommissionOffer],
) -> OrderLineItem:
    """Calculate commission and net income for one line inside an order."""
    line_gross = round(quantity * unit_price, 2)
    offer, match_method = lookup_offer_by_name(
        item_name, merchant_sku, offers_by_name, offers_by_sku
    )
    list_price = round(offer.price, 2) if offer and offer.price is not None else None
    commission_percent = offer.commission_percent if offer else None
    product_self_cost = unit_self_cost(offer)

    if commission_percent is None:
        net_income = calculate_net_income_totally(
            line_gross, 0.0, product_self_cost, quantity
        )
        return OrderLineItem(
            item_name=item_name,
            merchant_sku=merchant_sku or (offer.merchant_sku if offer else ""),
            quantity=quantity,
            line_gross=line_gross,
            list_price=list_price,
            commission_percent=None,
            commission_before_vat=0.0,
            commission_with_vat=0.0,
            commission_with_vat_per_item=0.0,
            product_self_cost=product_self_cost,
            net_income=net_income,
            net_income_per_item=net_income_per_unit(net_income, quantity),
            status="missing_commission" if offer else "not_found",
            match_method=match_method,
        )

    before_vat, with_vat = calculate_commission_amounts(line_gross, commission_percent)
    net_income = calculate_net_income_totally(
        line_gross, with_vat, product_self_cost, quantity
    )
    return OrderLineItem(
        item_name=item_name,
        merchant_sku=merchant_sku or (offer.merchant_sku if offer else ""),
        quantity=quantity,
        line_gross=line_gross,
        list_price=list_price,
        commission_percent=commission_percent,
        commission_before_vat=before_vat,
        commission_with_vat=with_vat,
        commission_with_vat_per_item=fee_per_item(with_vat, quantity),
        product_self_cost=product_self_cost,
        net_income=net_income,
        net_income_per_item=net_income_per_unit(net_income, quantity),
        status="ok",
        match_method=match_method,
    )


def calculate_orders_report(
    delivered_orders: list[dict[str, str]],
    offers_by_name: dict[str, CommissionOffer],
    offers_by_sku: dict[str, CommissionOffer],
    sku_map: dict[str, str],
) -> list[CalculatedOrder]:
    """Build per-order net income with expandable item lines (delivered only)."""
    calculated_orders: list[CalculatedOrder] = []

    for order in delivered_orders:
        order_number = (order.get("Order number") or "").strip()
        items_text = order.get("Items") or ""
        order_gross = float(str(order.get("Price", "0")).replace(",", "").strip())
        line_items: list[OrderLineItem] = []

        for item_match in ITEM_PATTERN.finditer(items_text):
            quantity = int(item_match.group(1))
            item_name = item_match.group(2).strip()
            unit_price = float(item_match.group(3))
            sku = sku_map.get(normalize_product_name(item_name), "")
            line_items.append(
                build_order_line_item(
                    item_name, quantity, unit_price, sku, offers_by_name, offers_by_sku
                )
            )

        if not line_items:
            line_items.append(
                OrderLineItem(
                    item_name=items_text,
                    merchant_sku="",
                    quantity=0,
                    line_gross=order_gross,
                    list_price=None,
                    commission_percent=None,
                    commission_before_vat=0.0,
                    commission_with_vat=0.0,
                    commission_with_vat_per_item=0.0,
                    product_self_cost=0.0,
                    net_income=calculate_net_income_totally(order_gross, 0.0, 0.0, 0),
                    net_income_per_item=0.0,
                    status="unparsed_items",
                    match_method="not_found",
                )
            )

        calculated_orders.append(
            CalculatedOrder(
                order_number=order_number,
                order_placed=(order.get("Order placed") or "").strip(),
                delivery_time=(order.get("Delivery time") or "").strip(),
                delivery_status=(order.get("Delivery status") or "").strip(),
                order_gross=round(order_gross, 2),
                commission_before_vat=round(sum(i.commission_before_vat for i in line_items), 2),
                commission_with_vat=round(sum(i.commission_with_vat for i in line_items), 2),
                net_income=round(sum(i.net_income for i in line_items), 2),
                items=line_items,
            )
        )

    calculated_orders.sort(key=lambda row: parse_wolt_datetime(row.delivery_time or row.order_placed))
    return calculated_orders


def parse_wolt_datetime(value: str) -> tuple[int, int, int, int, int]:
    """
    Parse Wolt CSV datetime like '15/06/2026, 20:39' for chronological sorting.
    Returns sortable tuple (year, month, day, hour, minute); unknown dates sort last.
    """
    match = re.match(r"(\d{2})/(\d{2})/(\d{4}), (\d{1,2}):(\d{2})", value or "")
    if not match:
        return (9999, 12, 31, 23, 59)
    return (
        int(match.group(3)),
        int(match.group(2)),
        int(match.group(1)),
        int(match.group(4)),
        int(match.group(5)),
    )


def parse_iso_date(value: str) -> date | None:
    """Parse YYYY-MM-DD from Wolt ad campaign labels."""
    try:
        year, month, day = value.split("-")
        return date(int(year), int(month), int(day))
    except (ValueError, AttributeError):
        return None


def parse_wolt_order_date(value: str) -> date | None:
    """Parse DD/MM/YYYY from orderNumbers Order placed / Delivery time."""
    match = re.match(r"(\d{2})/(\d{2})/(\d{4})", value or "")
    if not match:
        return None
    return date(int(match.group(3)), int(match.group(2)), int(match.group(1)))


def parse_ad_campaign_from_label(label: str, amount_incl_vat: float) -> AdCampaignCharge | None:
    """Extract attributed-purchase date window from a WOLT INVOICE ad campaign line."""
    if "Ad campaign" not in label:
        return None

    between = AD_CAMPAIGN_BETWEEN_PATTERN.search(label)
    if between:
        start = parse_iso_date(between.group(1))
        end = parse_iso_date(between.group(2))
        if start and end:
            return AdCampaignCharge(
                label=label,
                amount_incl_vat=round(amount_incl_vat, 2),
                start_date=start,
                end_date=end,
            )

    on_day = AD_CAMPAIGN_ON_PATTERN.search(label)
    if on_day:
        day = parse_iso_date(on_day.group(1))
        if day:
            return AdCampaignCharge(
                label=label,
                amount_incl_vat=round(amount_incl_vat, 2),
                start_date=day,
                end_date=day,
            )

    return None


def order_reference_date(order: CalculatedOrder) -> date | None:
    """Prefer order placed date for ad attribution windows (matches Wolt campaign labels)."""
    return parse_wolt_order_date(order.order_placed) or parse_wolt_order_date(order.delivery_time)


def distribute_amount_pro_rata(amount: float, weights: list[float]) -> list[float]:
    """Split a fee across rows proportional to weights; fix rounding on the last row."""
    if not weights or amount <= 0:
        return [0.0] * len(weights)

    total_weight = sum(weights)
    if total_weight <= 0:
        return [0.0] * len(weights)

    shares = [round(amount * weight / total_weight, 2) for weight in weights]
    rounding_delta = round(amount - sum(shares), 2)
    if shares and rounding_delta:
        shares[-1] = round(shares[-1] + rounding_delta, 2)
    return shares


def finalize_net_income_after_ad_cost(order: CalculatedOrder) -> None:
    """Ensure after-ad fields exist even when no ad allocation was applied."""
    order.allocated_ad_cost = round(order.allocated_ad_cost, 2)
    order.net_income_after_ad_cost = round(order.net_income - order.allocated_ad_cost, 2)
    for line in order.items:
        line.allocated_ad_cost = round(line.allocated_ad_cost, 2)
        line.net_income_after_ad_cost = round(line.net_income - line.allocated_ad_cost, 2)
        line.net_income_after_ad_cost_per_item = net_income_per_unit(
            line.net_income_after_ad_cost,
            line.quantity,
        )


def apply_ad_campaign_allocation(
    orders: list[CalculatedOrder],
    campaigns: list[AdCampaignCharge],
) -> dict[str, float]:
    """
    Allocate Wolt ad campaign fees to orders in each campaign's date window.

    Within a window, cost is split by order gross; within an order, by line gross.
    This is an estimate — Wolt bills attributed purchases, not every order in the range.
    """
    order_ad_totals: dict[str, float] = {order.order_number: 0.0 for order in orders}
    unallocated = 0.0

    for campaign in campaigns:
        eligible = [
            order
            for order in orders
            if (ref_date := order_reference_date(order)) is not None
            and campaign.start_date <= ref_date <= campaign.end_date
        ]
        shares = distribute_amount_pro_rata(
            campaign.amount_incl_vat,
            [order.order_gross for order in eligible],
        )
        if not eligible:
            unallocated = round(unallocated + campaign.amount_incl_vat, 2)
            continue

        for order, share in zip(eligible, shares):
            order_ad_totals[order.order_number] = round(
                order_ad_totals[order.order_number] + share,
                2,
            )

    allocated_total = 0.0
    for order in orders:
        order.allocated_ad_cost = round(order_ad_totals.get(order.order_number, 0.0), 2)
        allocated_total = round(allocated_total + order.allocated_ad_cost, 2)
        line_shares = distribute_amount_pro_rata(
            order.allocated_ad_cost,
            [line.line_gross for line in order.items],
        )
        for line, share in zip(order.items, line_shares):
            line.allocated_ad_cost = share

        finalize_net_income_after_ad_cost(order)

    return {
        "total_ad_campaigns_allocated_incl_vat": round(allocated_total, 2),
        "total_ad_campaigns_unallocated_incl_vat": round(unallocated, 2),
    }


def aggregate_products_from_orders(
    orders: list[CalculatedOrder],
) -> list[CalculatedRow]:
    """Roll up delivered order lines into per-product totals for the Products tab."""
    by_name: dict[str, CalculatedRow] = {}

    for order in orders:
        for line in order.items:
            key = normalize_product_name(line.item_name)
            if key not in by_name:
                by_name[key] = CalculatedRow(
                    item_name=line.item_name,
                    merchant_sku=line.merchant_sku,
                    quantity=0,
                    list_price=line.list_price,
                    list_total=None,
                    sold_total=0.0,
                    gross_total=0.0,
                    commission_percent=line.commission_percent,
                    commission_before_vat=0.0,
                    commission_with_vat=0.0,
                    commission_with_vat_per_item=0.0,
                    product_self_cost=line.product_self_cost,
                    net_income=0.0,
                    net_income_per_item=0.0,
                    status=line.status,
                    match_method=line.match_method,
                )

            row = by_name[key]
            row.quantity += line.quantity
            row.sold_total = round(row.sold_total + line.line_gross, 2)
            row.gross_total = row.sold_total
            row.commission_before_vat = round(row.commission_before_vat + line.commission_before_vat, 2)
            row.commission_with_vat = round(row.commission_with_vat + line.commission_with_vat, 2)
            row.net_income = round(row.net_income + line.net_income, 2)
            row.allocated_ad_cost = round(row.allocated_ad_cost + line.allocated_ad_cost, 2)
            row.net_income_after_ad_cost = round(
                row.net_income_after_ad_cost + line.net_income_after_ad_cost,
                2,
            )
            if line.product_self_cost:
                row.product_self_cost = line.product_self_cost
            if line.list_price is not None:
                row.list_price = line.list_price
            if line.commission_percent is not None:
                row.commission_percent = line.commission_percent
            if line.status != "ok":
                row.status = line.status

    rows = list(by_name.values())
    for row in rows:
        if row.list_price is not None:
            row.list_total = round(row.list_price * row.quantity, 2)
        row.net_income_per_item = net_income_per_unit(row.net_income, row.quantity)
        row.net_income_after_ad_cost_per_item = net_income_per_unit(
            row.net_income_after_ad_cost,
            row.quantity,
        )
        row.commission_with_vat_per_item = fee_per_item(row.commission_with_vat, row.quantity)

    rows.sort(key=lambda row: row.sold_total, reverse=True)
    return rows


def collect_missing_commission_products(
    rows: list[CalculatedRow],
) -> list[dict[str, Any]]:
    """
    Products sold in delivered orders with no commission % from offers_commission.xlsx.

    status missing_commission — matched in catalog but commission_home_delivery is empty.
    status not_found — no matching row in offers_commission.xlsx.
    """
    missing = [
        {
            "item_name": row.item_name,
            "merchant_sku": row.merchant_sku,
            "quantity": row.quantity,
            "sold_total": row.sold_total,
            "status": row.status,
            "match_method": row.match_method,
        }
        for row in rows
        if row.status in ("missing_commission", "not_found")
    ]
    missing.sort(key=lambda item: item["sold_total"], reverse=True)
    return missing


def build_summary_from_rows(
    rows: list[CalculatedRow],
    delivered_order_count: int = 0,
    rejected_order_count: int = 0,
    rejected_order_total: float = 0.0,
) -> CalculationSummary:
    """Compute period summary cards from calculated product rows."""
    matched_count = sum(1 for row in rows if row.status == "ok")
    total_sold = sum(row.sold_total for row in rows)
    total_list = sum(row.list_total if row.list_total is not None else row.sold_total for row in rows)

    return CalculationSummary(
        row_count=len(rows),
        matched_count=matched_count,
        unmatched_count=len(rows) - matched_count,
        delivered_order_count=delivered_order_count,
        rejected_order_count=rejected_order_count,
        rejected_order_total=round(rejected_order_total, 2),
        total_gross=round(total_sold, 2),
        total_list_value=round(total_list, 2),
        total_sold_value=round(total_sold, 2),
        total_commission_before_vat=round(sum(row.commission_before_vat for row in rows), 2),
        total_commission_with_vat=round(sum(row.commission_with_vat for row in rows), 2),
        total_net_income=round(sum(row.net_income for row in rows), 2),
        total_product_self_cost=round(
            sum(row.product_self_cost * row.quantity for row in rows), 2
        ),
    )


def aggregate_self_billing_expense_totals(
    lines: list[dict[str, Any]],
) -> dict[str, float]:
    """
    Sum SELF-BILLING rows (all except Total, goods sold) into expense adjustments.

    Negative row totals → added to expenses (sum of absolute values).
    Positive row totals → discounted from expenses.
    """
    negative_net = 0.0
    negative_incl = 0.0
    positive_net = 0.0
    positive_incl = 0.0

    for line in lines:
        total = float(line["amount"])
        net = float(line["net"])
        if total < 0:
            negative_net += abs(net)
            negative_incl += abs(total)
        elif total > 0:
            positive_net += net
            positive_incl += total

    net_add_net = round(negative_net - positive_net, 2)
    net_add_incl = round(negative_incl - positive_incl, 2)

    return {
        "self_billing_negative_sum_net": round(negative_net, 2),
        "self_billing_negative_sum_incl_vat": round(negative_incl, 2),
        "self_billing_positive_sum_net": round(positive_net, 2),
        "self_billing_positive_sum_incl_vat": round(positive_incl, 2),
        "total_self_billing_deductions_net": net_add_net,
        "total_self_billing_deductions_incl_vat": net_add_incl,
    }


def self_billing_expense_contribution(net: float, total_incl_vat: float) -> tuple[float, float]:
    """
    Map a SELF-BILLING INVOICE row (except Total, goods sold) to expense deltas.

    Wolt convention in standardSummary:
        negative TOTAL → add to merchant expenses (e.g. Remunerations -610.04)
        positive TOTAL → reduce expenses (credit / discount to merchant)
    """
    if total_incl_vat < 0:
        return abs(net), abs(total_incl_vat)
    if total_incl_vat > 0:
        return -abs(net), -abs(total_incl_vat)
    return 0.0, 0.0


def parse_payment_details_csv(csv_text: str) -> dict[str, Any]:
    """
    Extract self-billing and WOLT INVOICE totals from standardSummary.csv
    (or legacy payment_details.csv).

    Column layout (semicolon-separated):
        …; label; NET; VAT; TOTAL (incl. VAT)

    Total Wolt expenses = WOLT INVOICE charges + SELF-BILLING adjustments:
        Every data row between SELF-BILLING INVOICE and END PAYOUT is included
        except "Total, goods sold" (revenue only).
        Negative row totals add to expenses; positive row totals reduce expenses.
    """
    result: dict[str, Any] = {
        "wolt_invoice_lines": [],
        "self_billing_lines": [],
    }
    section: str | None = None
    wolt_net_total = 0.0
    wolt_vat_total = 0.0
    distribution_incl_vat = 0.0

    for line in csv_text.splitlines():
        if "Date of issue" in line and "WOLT INVOICE" in line:
            section = "wolt"
            continue
        if "Date of issue" in line and "SELF-BILLING INVOICE" in line:
            section = "self_billing"
            continue
        if "END PAYOUT" in line or "Payout amount" in line:
            section = "payout"

        parts = [p.strip() for p in line.split(";")]
        if len(parts) < 5:
            continue

        row_label = parts[4] if len(parts) > 4 else ""
        if not row_label or row_label in ("WOLT INVOICE", "SELF-BILLING INVOICE"):
            continue

        last_value = parts[-1].replace(",", "").strip()
        try:
            amount = float(last_value)
        except ValueError:
            continue

        if row_label == "Payout amount":
            result["payout_amount"] = amount
            continue

        if row_label == "Total, goods sold":
            result["gross_goods_sold"] = amount
            continue

        # Parse NET / VAT / TOTAL when present (standardSummary data rows).
        net = vat = total = amount
        if len(parts) >= 8:
            try:
                net = float(parts[5].replace(",", ""))
                vat = float(parts[6].replace(",", ""))
                total = float(parts[7].replace(",", ""))
            except ValueError:
                pass

        if section == "wolt":
            result["wolt_invoice_lines"].append(
                {
                    "label": row_label,
                    "net": round(net, 2),
                    "vat": round(vat, 2),
                    "amount": round(total, 2),
                }
            )
            wolt_net_total += net
            wolt_vat_total += vat
            if row_label.startswith("Distribution,"):
                distribution_incl_vat += total
            continue

        if section == "self_billing":
            result["self_billing_lines"].append(
                {
                    "label": row_label,
                    "net": round(net, 2),
                    "vat": round(vat, 2),
                    "amount": round(total, 2),
                }
            )
            if row_label == "Total, goods discounted by merchant":
                result["merchant_discounts"] = abs(total)
            elif row_label == "Remunerations" and total < 0:
                result["remunerations"] = abs(total)

    if result["wolt_invoice_lines"]:
        result["total_wolt_invoice"] = round(
            sum(line["amount"] for line in result["wolt_invoice_lines"]), 2
        )
        result["total_wolt_invoice_net"] = round(wolt_net_total, 2)
        result["total_wolt_invoice_vat"] = round(wolt_vat_total, 2)
        result["total_wolt_distribution_incl_vat"] = round(distribution_incl_vat, 2)

        ad_campaigns: list[AdCampaignCharge] = []
        ad_total = 0.0
        other_fees = 0.0
        for line in result["wolt_invoice_lines"]:
            label = line["label"]
            amount = float(line["amount"])
            if label.startswith("Distribution,"):
                continue
            campaign = parse_ad_campaign_from_label(label, amount)
            if campaign:
                ad_campaigns.append(campaign)
                ad_total += amount
            else:
                other_fees += amount
        result["ad_campaigns"] = ad_campaigns
        result["total_ad_campaigns_incl_vat"] = round(ad_total, 2)
        result["total_other_fees_incl_vat"] = round(other_fees, 2)

    if result["self_billing_lines"]:
        sb_totals = aggregate_self_billing_expense_totals(result["self_billing_lines"])
        result.update(sb_totals)

    wolt_invoice_incl = result.get("total_wolt_invoice", 0.0)
    wolt_invoice_net = result.get("total_wolt_invoice_net", 0.0)
    self_billing_add_incl = result.get("total_self_billing_deductions_incl_vat", 0.0)
    self_billing_add_net = result.get("total_self_billing_deductions_net", 0.0)
    result["total_wolt_expenses_incl_vat"] = round(
        wolt_invoice_incl + self_billing_add_incl, 2
    )
    result["total_wolt_expenses_net"] = round(wolt_invoice_net + self_billing_add_net, 2)

    return result


def enrich_summary_with_wolt_summary(
    summary: CalculationSummary,
    payment_meta: dict[str, Any] | None,
) -> CalculationSummary:
    """Overlay Wolt standardSummary totals onto Financial snapshot KPIs."""
    if not payment_meta or payment_meta.get("gross_goods_sold") is None:
        return summary

    payout = payment_meta.get("payout_amount")
    wolt_net_income = (
        round(payout - summary.total_product_self_cost, 2)
        if payout is not None
        else None
    )

    return CalculationSummary(
        row_count=summary.row_count,
        matched_count=summary.matched_count,
        unmatched_count=summary.unmatched_count,
        delivered_order_count=summary.delivered_order_count,
        rejected_order_count=summary.rejected_order_count,
        rejected_order_total=summary.rejected_order_total,
        total_gross=summary.total_gross,
        total_list_value=summary.total_list_value,
        total_sold_value=summary.total_sold_value,
        total_commission_before_vat=summary.total_commission_before_vat,
        total_commission_with_vat=summary.total_commission_with_vat,
        total_net_income=summary.total_net_income,
        total_product_self_cost=summary.total_product_self_cost,
        wolt_summary_gross_goods=payment_meta.get("gross_goods_sold"),
        wolt_summary_expenses_net=payment_meta.get("total_wolt_expenses_net"),
        wolt_summary_expenses_incl_vat=payment_meta.get("total_wolt_expenses_incl_vat"),
        wolt_summary_distribution_incl_vat=payment_meta.get("total_wolt_distribution_incl_vat"),
        wolt_summary_remunerations=payment_meta.get("remunerations"),
        wolt_summary_self_billing_deductions_incl_vat=payment_meta.get(
            "total_self_billing_deductions_incl_vat"
        ),
        wolt_summary_self_billing_negative_incl_vat=payment_meta.get(
            "self_billing_negative_sum_incl_vat"
        ),
        wolt_summary_payout=payment_meta.get("payout_amount"),
        wolt_summary_net_income=wolt_net_income,
    )


def enrich_summary_with_expense_breakdown(
    summary: CalculationSummary,
    payment_meta: dict[str, Any] | None,
    ad_allocation: dict[str, float] | None = None,
) -> CalculationSummary:
    """Explain which Wolt invoice expenses are excluded from default per-item net income."""
    if not payment_meta or payment_meta.get("gross_goods_sold") is None:
        return summary

    wolt_expenses = float(payment_meta.get("total_wolt_expenses_incl_vat") or 0.0)
    distribution = float(payment_meta.get("total_wolt_distribution_incl_vat") or 0.0)
    excluded_total = round(wolt_expenses - summary.total_commission_with_vat, 2)
    distribution_gap = round(distribution - summary.total_commission_with_vat, 2)
    ad_allocated = float(
        (ad_allocation or {}).get("total_ad_campaigns_allocated_incl_vat") or 0.0
    )
    excluded_after_ads = round(excluded_total - ad_allocated, 2)

    return replace(
        summary,
        wolt_summary_ad_campaigns_incl_vat=payment_meta.get("total_ad_campaigns_incl_vat"),
        wolt_summary_ad_campaigns_allocated_incl_vat=round(ad_allocated, 2),
        wolt_summary_other_fees_incl_vat=payment_meta.get("total_other_fees_incl_vat"),
        wolt_summary_distribution_gap_incl_vat=distribution_gap,
        per_item_expenses_excluded_incl_vat=excluded_total,
        per_item_expenses_excluded_after_ads_incl_vat=excluded_after_ads,
    )


def build_invoice_reconciliation(
    summary: CalculationSummary,
    payment_details: dict[str, float] | None,
) -> InvoiceReconciliation:
    """
    Build a step-by-step invoice waterfall for the UI.

    Per-item net income is unchanged — this only explains how Wolt invoice lines
    relate to the delivered-order total (commission base).
    """
    net_sold_orders = round(summary.total_sold_value, 2)
    wolt_fees = round(summary.total_commission_with_vat, 2)
    net_after_wolt = round(summary.total_net_income, 2)

    gross = payment_details.get("gross_goods_sold") if payment_details else None
    merchant_disc = payment_details.get("merchant_discounts") if payment_details else None
    remunerations = payment_details.get("remunerations") if payment_details else None
    payout = payment_details.get("payout_amount") if payment_details else None

    net_from_invoice = (
        round(gross - merchant_disc, 2)
        if gross is not None and merchant_disc is not None
        else None
    )
    orders_match = (
        abs(net_from_invoice - net_sold_orders) < 0.02 if net_from_invoice is not None else None
    )
    net_after_remunerations = (
        round(net_sold_orders - remunerations, 2) if remunerations is not None else None
    )

    steps: list[InvoiceStep] = []
    phase1_steps: list[InvoiceStep] = []
    phase2_steps: list[InvoiceStep] = []
    phase3_steps: list[InvoiceStep] = []
    running = 0.0

    if gross is not None:
        running = gross
        phase1_steps.append(
            InvoiceStep(
                id="gross_goods",
                label="Gross goods sold (Wolt self-billing invoice)",
                label_he="סה״כ מכירות כולל מע״מ",
                amount=gross,
                running_total=running,
                step_type="start",
                note="From MERCHANT_TO_WOLT invoice — menu/list sales before your discounts",
                phase="1",
            )
        )

    if merchant_disc is not None and gross is not None:
        running = round(running - merchant_disc, 2)
        phase1_steps.append(
            InvoiceStep(
                id="merchant_discounts",
                label="Merchant discounts",
                label_he="סך הנחות בית העסק",
                amount=-merchant_disc,
                running_total=running,
                step_type="subtract",
                note="Includes per-item reductions (e.g. ₪5,250 → ₪5,225 on iPhone). Already reflected in order prices.",
                phase="1",
            )
        )

    phase1_steps.append(
        InvoiceStep(
            id="net_sold_orders",
            label="Net sold — delivered orders (commission base)",
            label_he="מכירות נטו מההזמנות שסופקו",
            amount=net_sold_orders,
            running_total=net_sold_orders,
            step_type="milestone",
            note="Sum of delivered orderNumbers. Per-item commission is calculated on this — unchanged.",
            phase="1",
        )
    )

    if remunerations is not None:
        running = round(net_sold_orders - remunerations, 2)
        phase1_steps.append(
            InvoiceStep(
                id="remunerations",
                label="Remunerations / deductions",
                label_he="סה״כ ניכויים",
                amount=-remunerations,
                running_total=running,
                step_type="subtract",
                note="Period-level Wolt adjustment. Display only — does not change per-item net income.",
                phase="1",
            )
        )

    phase2_steps.append(
        InvoiceStep(
            id="wolt_fees",
            label="Wolt distribution fees (incl. VAT ×1.18)",
            label_he="עמלת Wolt כולל מע״מ",
            amount=-wolt_fees,
            running_total=round(net_sold_orders - wolt_fees, 2),
            step_type="subtract",
            note="Sum of per-item fees from commission %. Your existing net income calculation.",
            phase="2",
        )
    )

    phase2_steps.append(
        InvoiceStep(
            id="net_income",
            label="Net income after Wolt fees",
            label_he="הכנסה נטו לאחר עמלת Wolt",
            amount=net_after_wolt,
            running_total=net_after_wolt,
            step_type="result",
            note="Final per-item / per-order net income shown in the tables below.",
            phase="2",
        )
    )

    total_wolt_invoice = None
    payout_gap = None
    if payment_details:
        total_wolt_invoice = payment_details.get("total_wolt_invoice")
        wolt_lines = payment_details.get("wolt_invoice_lines", [])

        if net_after_remunerations is not None and payout is not None:
            payout_gap = round(net_after_wolt - payout, 2)
            phase3_steps.append(
                InvoiceStep(
                    id="app_net_reference",
                    label="App net income (distribution fees only)",
                    label_he="הכנסה נטו באפליקציה",
                    amount=net_after_wolt,
                    running_total=net_after_wolt,
                    step_type="milestone",
                    note="What the dashboard calculates — product sales minus distribution commission only.",
                    phase="3",
                )
            )
            phase3_steps.append(
                InvoiceStep(
                    id="payout_gap_note",
                    label="Gap to bank payout",
                    label_he="הפער לתשלום בפועל",
                    amount=-payout_gap,
                    running_total=payout_gap if payout_gap is not None else 0,
                    step_type="subtract",
                    note="Difference is ads, lateness, delivery discount, remunerations, and full Wolt invoice vs app fees.",
                    phase="3",
                )
            )

            phase3_steps.append(
                InvoiceStep(
                    id="payout_start",
                    label="Net goods after remunerations",
                    label_he="נטו לאחר ניכויים",
                    amount=net_after_remunerations,
                    running_total=net_after_remunerations,
                    step_type="milestone",
                    note="Starting point for Wolt bank payout calculation.",
                    phase="3",
                )
            )

            running_payout = net_after_remunerations
            for wolt_line in wolt_lines:
                running_payout = round(running_payout - wolt_line["amount"], 2)
                phase3_steps.append(
                    InvoiceStep(
                        id=f"wolt_{len(phase3_steps)}",
                        label=wolt_line["label"],
                        label_he="חיוב Wolt",
                        amount=-wolt_line["amount"],
                        running_total=running_payout,
                        step_type="subtract",
                        note="From WOLT INVOICE section in payment_details.csv",
                        phase="3",
                    )
                )

            if total_wolt_invoice is not None and net_after_remunerations is not None:
                phase3_steps.append(
                    InvoiceStep(
                        id="total_wolt_invoice",
                        label="Total WOLT INVOICE (all period charges)",
                        label_he="סה״כ חשבונית Wolt",
                        amount=-total_wolt_invoice,
                        running_total=round(net_after_remunerations - total_wolt_invoice, 2),
                        step_type="subtract",
                        note="Includes distribution, ads, lateness, delivery discount, and VAT adjustments.",
                        phase="3",
                    )
                )

            if payout is not None:
                phase3_steps.append(
                    InvoiceStep(
                        id="payout",
                        label="Bank payout (actual transfer)",
                        label_he="סכום התשלום לחשבון הבנק",
                        amount=payout,
                        running_total=payout,
                        step_type="result",
                        note="What Wolt deposits — after all invoice lines in the period.",
                        phase="3",
                    )
                )

    steps = phase1_steps + phase2_steps + phase3_steps
    phases = [
        InvoicePhase(
            id="1",
            title="Wolt invoice → net goods",
            subtitle="Gross sales, merchant discounts, remunerations",
            steps=phase1_steps,
        ),
        InvoicePhase(
            id="2",
            title="Commission on orders",
            subtitle="Distribution fees — per-item calculation (unchanged)",
            steps=phase2_steps,
        ),
    ]
    if phase3_steps:
        phases.append(
            InvoicePhase(
                id="3",
                title="Why payout ≠ app net income",
                subtitle="Full WOLT INVOICE charges → bank transfer",
                steps=phase3_steps,
            )
        )

    source = "payment_details.csv" if payment_details else "orders_only"
    return InvoiceReconciliation(
        source=source,
        gross_goods_sold=gross,
        merchant_discounts=merchant_disc,
        net_sold_from_invoice=net_from_invoice,
        net_sold_from_orders=net_sold_orders,
        orders_match_invoice=orders_match,
        remunerations=remunerations,
        net_after_remunerations=net_after_remunerations,
        wolt_distribution_fees=wolt_fees,
        net_income_after_wolt=net_after_wolt,
        payout_amount=payout,
        total_wolt_invoice=total_wolt_invoice,
        payout_gap_from_app_net=payout_gap,
        steps=steps,
        phases=phases,
    )


def run_calculation(
    offers_path: Path,
    order_numbers_csv: str | None = None,
    items_sold_csv: str | None = None,
    payment_details_csv: str | None = None,
    legacy_items_sold_csv: str | None = None,
) -> dict[str, Any]:
    """
    End-to-end calculation used by the serverless HTTP handler.

    Preferred: orderNumbers.csv (delivered orders, excludes rejected) + optional itemsSold for SKUs.
    Legacy: itemsSold.csv only (includes rejected quantities — not invoice-accurate).
    """
    offers_by_name = load_offers_from_xlsx(offers_path)
    offers_by_name, offers_by_sku = build_offer_indexes(offers_by_name)

    formula = {
        "commission_base": "line_gross from delivered orderNumbers (actual incl. VAT, per Wolt invoice)",
        "commission_before_vat": "sold_total * (commission_percent / 100)",
        "commission_with_vat": "commission_before_vat * 1.18",
        "net_income": "sold_total - commission_with_vat - (product_self_cost * quantity)",
    }

    # --- Primary path: orderNumbers (invoice-conformant) ---
    if order_numbers_csv:
        delivered, rejected = parse_order_numbers_csv(order_numbers_csv)
        rejected_total = sum(
            float(str(o.get("Price", "0")).replace(",", "").strip()) for o in rejected
        )
        sku_map: dict[str, str] = {}
        if items_sold_csv:
            sku_map = build_sku_map_from_items_sold(parse_items_sold_csv(items_sold_csv))

        orders = calculate_orders_report(delivered, offers_by_name, offers_by_sku, sku_map)

        payment_meta = (
            parse_payment_details_csv(payment_details_csv) if payment_details_csv else None
        )
        ad_allocation: dict[str, float] = {}
        if payment_meta and payment_meta.get("ad_campaigns"):
            ad_allocation = apply_ad_campaign_allocation(
                orders,
                payment_meta["ad_campaigns"],
            )
        else:
            for order in orders:
                finalize_net_income_after_ad_cost(order)

        rows = aggregate_products_from_orders(orders)
        summary = build_summary_from_rows(
            rows,
            delivered_order_count=len(delivered),
            rejected_order_count=len(rejected),
            rejected_order_total=rejected_total,
        )

        summary = enrich_summary_with_wolt_summary(summary, payment_meta)
        summary = enrich_summary_with_expense_breakdown(summary, payment_meta, ad_allocation)
        invoice = build_invoice_reconciliation(summary, payment_meta)

        invoice_dict = asdict(invoice)
        invoice_dict["steps"] = [asdict(step) for step in invoice.steps]
        invoice_dict["phases"] = [
            {**asdict(phase), "steps": [asdict(step) for step in phase.steps]}
            for phase in invoice.phases
        ]

        return {
            "summary": asdict(summary),
            "rows": [asdict(row) for row in rows],
            "orders": [asdict(order) for order in orders],
            "missing_commission_products": collect_missing_commission_products(rows),
            "invoice_reconciliation": invoice_dict,
            "data_source": "orderNumbers_delivered_only",
            "rejected_excluded": True,
            "upload_format": "orderNumbers.csv (required) + standardSummary.csv (optional) + itemsSold.csv (optional)",
            "formula": formula,
        }

    # --- Legacy path: itemsSold only ---
    csv_text = items_sold_csv or legacy_items_sold_csv
    if not csv_text:
        raise ValueError(
            "Upload orderNumbers.csv (required for invoice-accurate results). "
            "itemsSold.csv is optional for merchant SKUs."
        )

    sales_rows = parse_items_sold_csv(csv_text)
    calculated_rows, legacy_summary = calculate_net_income_report(
        sales_rows, offers_by_name, offers_by_sku
    )
    summary_dict = asdict(legacy_summary)
    summary_dict.update(
        {
            "delivered_order_count": 0,
            "rejected_order_count": 0,
            "rejected_order_total": 0.0,
        }
    )

    return {
        "summary": summary_dict,
        "rows": [asdict(row) for row in calculated_rows],
        "orders": [],
        "missing_commission_products": collect_missing_commission_products(calculated_rows),
        "data_source": "itemsSold_only",
        "rejected_excluded": False,
        "warning": "itemsSold includes rejected orders. Upload orderNumbers.csv for invoice-accurate totals.",
        "upload_format": "Wolt itemsSold CSV (Item name, Total, Quantity, Merchant SKU)",
        "formula": formula,
    }
